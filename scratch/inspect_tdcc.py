import os
import openpyxl
import pandas as pd

excel_path = "/Users/jeffrey/Downloads/集保股權集中TDCC(Buy).xlsx"

print("Exists:", os.path.exists(excel_path))
if os.path.exists(excel_path):
    try:
        # Load workbook using openpyxl just to inspect sheets
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        print("Sheet names:", wb.sheetnames)
        wb.close()
        
        # Load first few rows of each sheet using pandas
        for sheet in wb.sheetnames[:5]: # inspect first 5 sheets
            print(f"\n--- Sheet: {sheet} ---")
            df = pd.read_excel(excel_path, sheet_name=sheet, nrows=10)
            print("Columns:", list(df.columns))
            print("Shape:", df.shape)
            print(df.head(5))
    except Exception as e:
        print("Error during inspection:", e)
